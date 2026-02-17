import re
import pandas as pd
from ortools.sat.python import cp_model

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

INPUT_FILE = "Plantilla_Clubes_REAL.xlsx"
OUTPUT_FILE = "Fixture_2026.xlsx"
SHEET_CLUBES = "CLUBES"


# -------------------------
# Utils
# -------------------------
def as_bool(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return False
    if isinstance(v, str):
        t = v.strip().lower()
        return t in ("x", "1", "si", "sí", "s", "true", "ok")
    try:
        return int(v) == 1
    except Exception:
        return False


def normalize_team_series(s: pd.Series) -> pd.Series:
    return (
        s.astype(str)
        .str.replace("\u00a0", " ", regex=False)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
    )


def normalize_cat_name(name: str) -> str:
    # normaliza para comparar/deduplicar
    return re.sub(r"\s+", " ", str(name).strip()).upper()


def normalized_pairs(team_list):
    pairs = set()
    for i in range(len(team_list)):
        for j in range(i + 1, len(team_list)):
            a, b = sorted((team_list[i], team_list[j]))
            pairs.add((a, b))
    return sorted(pairs)


def rounds_for_n(n: int) -> int:
    return (n - 1) if (n % 2 == 0) else n


def safe_sheet_name(name: str) -> str:
    # Excel limita a 31 chars y prohíbe ciertos caracteres
    bad = ['\\', '/', '*', '[', ']', ':', '?']
    out = "".join("_" if ch in bad else ch for ch in str(name))
    out = out.strip()
    return out[:31] if len(out) > 31 else out


def detect_categories(df: pd.DataFrame):
    """
    Detecta categorías dinámicamente:
    - Cualquier columna que NO sea N°/EQUIPO
    - y que tenga al menos un valor truthy (X/1/si/ok)
    - deduplicadas por nombre normalizado (evita S y 'S ')
    """
    reserved = {"EQUIPO", "N°", "Nº", "NRO", "NRO.", "NUMERO", "NÚMERO"}
    seen = set()
    cats = []

    for col in df.columns:
        col_clean = str(col).strip()
        if normalize_cat_name(col_clean) in reserved:
            continue

        if df[col].apply(as_bool).any():
            key = normalize_cat_name(col_clean)
            if key not in seen:
                seen.add(key)
                cats.append(col_clean)  # conservamos el nombre original “humano”

    return cats


# -------------------------
# Solver (dinámico)
# -------------------------
def solve_fixture(df: pd.DataFrame, cats: list[str]):
    if "EQUIPO" not in df.columns:
        raise ValueError(f"Falta la columna 'EQUIPO'. Columnas: {list(df.columns)}")

    df = df.copy()
    df["EQUIPO"] = normalize_team_series(df["EQUIPO"])

    all_teams = sorted(df["EQUIPO"].dropna().unique().tolist())

    teams_by_cat = {}
    for c in cats:
        if c not in df.columns:
            raise ValueError(f"Falta la columna '{c}' en la plantilla.")
        teams_by_cat[c] = sorted(df[df[c].apply(as_bool)]["EQUIPO"].dropna().unique().tolist())

    for c in cats:
        if len(teams_by_cat[c]) < 2:
            raise ValueError(f"La categoría '{c}' tiene menos de 2 equipos (n={len(teams_by_cat[c])}).")

    rounds_by_cat = {c: rounds_for_n(len(teams_by_cat[c])) for c in cats}
    R = max(rounds_by_cat.values())  # fechas totales

    model = cp_model.CpModel()

    x = {}    # x[(cat, r, a, b)]
    bye = {}  # bye[(cat, r, team)]
    pairs_by_cat = {c: normalized_pairs(teams_by_cat[c]) for c in cats}

    # Variables
    for c in cats:
        tlist = teams_by_cat[c]
        n = len(tlist)
        for r in range(R):
            for (a, b) in pairs_by_cat[c]:
                x[(c, r, a, b)] = model.NewBoolVar(f"x_{normalize_cat_name(c)}_{r}_{a}_{b}")
            if n % 2 == 1:
                for t in tlist:
                    bye[(c, r, t)] = model.NewBoolVar(f"bye_{normalize_cat_name(c)}_{r}_{t}")

    # Restricciones round-robin por categoría
    for c in cats:
        tlist = teams_by_cat[c]
        n = len(tlist)
        real_rounds = rounds_by_cat[c]

        # Fechas sobrantes: categoría inactiva
        for r in range(real_rounds, R):
            for (a, b) in pairs_by_cat[c]:
                model.Add(x[(c, r, a, b)] == 0)
            if n % 2 == 1:
                for t in tlist:
                    model.Add(bye[(c, r, t)] == 0)

        # Fechas reales: cada equipo juega 1 o bye
        for r in range(real_rounds):
            for t in tlist:
                incident = []
                for u in tlist:
                    if t == u:
                        continue
                    a, b = sorted((t, u))
                    incident.append(x[(c, r, a, b)])

                if n % 2 == 0:
                    model.Add(sum(incident) == 1)
                else:
                    model.Add(sum(incident) + bye[(c, r, t)] == 1)

            if n % 2 == 1:
                model.Add(sum(bye[(c, r, t)] for t in tlist) == 1)

        # ✅ todos contra todos: cada par exactamente 1 vez
        for (a, b) in pairs_by_cat[c]:
            model.Add(sum(x[(c, r, a, b)] for r in range(real_rounds)) == 1)

    # Objetivo: maximizar coincidencias por fecha entre TODAS las categorías
    weight_match = 50
    rewards = []
    for r in range(R):
        for i in range(len(cats)):
            for j in range(i + 1, len(cats)):
                c1, c2 = cats[i], cats[j]
                common_pairs = sorted(set(pairs_by_cat[c1]) & set(pairs_by_cat[c2]))
                for (a, b) in common_pairs:
                    y = model.NewBoolVar(f"y_{normalize_cat_name(c1)}_{normalize_cat_name(c2)}_{r}_{a}_{b}")
                    model.Add(y <= x[(c1, r, a, b)])
                    model.Add(y <= x[(c2, r, a, b)])
                    model.Add(y >= x[(c1, r, a, b)] + x[(c2, r, a, b)] - 1)
                    rewards.append(weight_match * y)

    model.Maximize(sum(rewards))

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 25.0
    solver.parameters.num_search_workers = 8

    status = solver.Solve(model)
    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        raise RuntimeError("No se encontró un fixture viable con las restricciones.")

    # Tablas por categoría
    cat_tables = {}
    for c in cats:
        real_rounds = rounds_by_cat[c]
        rows = []
        for r in range(real_rounds):
            for (a, b) in pairs_by_cat[c]:
                if solver.Value(x[(c, r, a, b)]) == 1:
                    rows.append({"Fecha": r + 1, "Local": a, "Visitante": b})
            if len(teams_by_cat[c]) % 2 == 1:
                for t in teams_by_cat[c]:
                    if solver.Value(bye[(c, r, t)]) == 1:
                        rows.append({"Fecha": r + 1, "Local": t, "Visitante": "LIBRE"})

        cat_tables[c] = (
            pd.DataFrame(rows)
            .sort_values(["Fecha", "Local"])
            .reset_index(drop=True)
        )

    # Calendario unificado
    maps = {c: {} for c in cats}
    for c in cats:
        for _, row in cat_tables[c].iterrows():
            f = int(row["Fecha"])
            a = row["Local"]
            b = row["Visitante"]
            maps[c].setdefault(f, {})
            if b == "LIBRE":
                maps[c][f][a] = "LIBRE"
            else:
                maps[c][f][a] = b
                maps[c][f][b] = a

    cal_rows = []
    for f in range(1, R + 1):
        for t in all_teams:
            rivals = {c: maps[c].get(f, {}).get(t, "—") for c in cats}
            real = [rivals[c] for c in cats if rivals[c] not in ("—", "LIBRE")]
            ok = "✅" if len(set(real)) <= 1 else "⚠️"

            row = {"Fecha": f, "Equipo": t}
            for c in cats:
                row[f"Rival {c}"] = rivals[c]
            row["OK?"] = ok
            cal_rows.append(row)

    calendario = pd.DataFrame(cal_rows)
    return cat_tables, calendario, rounds_by_cat, R


# -------------------------
# Excel styling
# -------------------------
def style_table(ws, title: str):
    header_font = Font(bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    fill_title = PatternFill("solid", fgColor="111827")
    fill_header = PatternFill("solid", fgColor="F3F4F6")

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title row
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    t = ws.cell(row=1, column=1)
    t.value = title
    t.font = Font(bold=True, size=14, color="FFFFFF")
    t.fill = fill_title
    t.alignment = center
    ws.row_dimensions[1].height = 26

    # Header row
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=2, column=col)
        cell.font = header_font
        cell.fill = fill_header
        cell.alignment = center if col <= 2 else left
        cell.border = border

        # ancho base
        ws.column_dimensions[get_column_letter(col)].width = 22

    ws.freeze_panes = "A3"

    # Borders body
    for r in range(3, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).border = border


def style_info_sheet(ws):
    """
    INFO tiene 2 bloques: resumen (fila 2) y tabla por categoría (desde fila 6).
    Le ponemos un título principal + un subtítulo "Totales por categoría".
    """
    title = "INFO DE FECHAS / CATEGORÍAS"

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    fill_title = PatternFill("solid", fgColor="111827")
    fill_header = PatternFill("solid", fgColor="F3F4F6")

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Insert title row
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    t = ws.cell(row=1, column=1)
    t.value = title
    t.font = Font(bold=True, size=14, color="FFFFFF")
    t.fill = fill_title
    t.alignment = center
    ws.row_dimensions[1].height = 26

    # Estilo encabezados del resumen (fila 2)
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=2, column=col)
        cell.font = Font(bold=True, size=11)
        cell.fill = fill_header
        cell.alignment = left
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = 22

    # Subtítulo (fila 5): "Totales por categoría"
    ws["A5"].value = "Totales por categoría"
    ws["A5"].font = Font(bold=True, size=12)
    ws["A5"].alignment = left

    # Headers de tabla categorías (fila 6)
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=6, column=col)
        if cell.value is None:
            continue
        cell.font = Font(bold=True, size=11)
        cell.fill = fill_header
        cell.alignment = left
        cell.border = border

    # Bordes para el cuerpo (desde fila 3 en adelante)
    for r in range(3, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).border = border

    ws.freeze_panes = "A3"


def write_partidos_por_fecha_pretty(ws, cat_tables: dict, cats: list[str]):
    title_font = Font(bold=True, size=14, color="FFFFFF")
    fecha_font = Font(bold=True, size=12)
    header_font = Font(bold=True, size=11)

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    fill_title = PatternFill("solid", fgColor="111827")
    fill_fecha = PatternFill("solid", fgColor="E5E7EB")
    fill_header = PatternFill("solid", fgColor="F3F4F6")

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    rows = []
    for cat in cats:
        dfc = cat_tables.get(cat)
        if dfc is None or dfc.empty:
            continue
        for _, r in dfc.iterrows():
            rows.append((int(r["Fecha"]), cat, str(r["Local"]), str(r["Visitante"])))

    rows.sort(key=lambda x: (x[0], x[1], x[2], x[3]))

    ws.append(["PARTIDOS POR FECHA", "", "", ""])
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.font = title_font
    c.alignment = center
    c.fill = fill_title
    ws.row_dimensions[1].height = 26

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 30

    current_row = 2
    current_fecha = None

    def write_headers():
        nonlocal current_row
        ws.append(["Categoría", "Cancha", "Local", "Visitante"])
        for col in range(1, 5):
            cell = ws.cell(row=current_row, column=col)
            cell.font = header_font
            cell.fill = fill_header
            cell.alignment = center if col <= 2 else left
            cell.border = border
        ws.row_dimensions[current_row].height = 18
        current_row += 1

    for fecha, cat, local, visitante in rows:
        if fecha != current_fecha:
            if current_fecha is not None:
                ws.append(["", "", "", ""])
                current_row += 1

            ws.append([f"FECHA {fecha}", "", "", ""])
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
            cell = ws.cell(row=current_row, column=1)
            cell.font = fecha_font
            cell.fill = fill_fecha
            cell.alignment = left
            ws.row_dimensions[current_row].height = 20
            current_row += 1

            write_headers()
            current_fecha = fecha

        ws.append([cat, "", local, visitante])
        for col in range(1, 5):
            cell = ws.cell(row=current_row, column=col)
            cell.alignment = center if col <= 2 else left
            cell.border = border
        ws.row_dimensions[current_row].height = 18
        current_row += 1

    ws.freeze_panes = "A2"


# -------------------------
# Main
# -------------------------
def main():
    df = pd.read_excel(INPUT_FILE, sheet_name=SHEET_CLUBES, engine="openpyxl")

    if "EQUIPO" not in df.columns:
        raise ValueError(f"Falta la columna 'EQUIPO'. Columnas: {list(df.columns)}")

    # Categorías dinámicas
    cats = detect_categories(df)
    if not cats:
        raise ValueError("No se detectó ninguna categoría. Agregá columnas con X/1/si/ok (ej: S, SS, M...).")

    # Resolver
    cat_tables, calendario, rounds_by_cat, R = solve_fixture(df, cats)

    # Armar INFO en 2 bloques (sin repetición)
    df_norm = df.copy()
    df_norm["EQUIPO"] = normalize_team_series(df_norm["EQUIPO"])

    info_resumen = pd.DataFrame([{
        "Fechas_totales": int(R),
        "Cantidad_categorias": int(len(cats)),
        "Categorias_detectadas": ", ".join(cats),
        "Cantidad_clubes_total": int(df_norm["EQUIPO"].dropna().nunique()),
    }])

    info_cats_rows = []
    for c in cats:
        equipos_c = int(df_norm[df_norm[c].apply(as_bool)]["EQUIPO"].dropna().nunique())
        fechas_c = int(rounds_by_cat[c])
        info_cats_rows.append({
            "Categoria": c,
            "Equipos": equipos_c,
            "Fechas": fechas_c,
            "Byes_por_fecha": 1 if (equipos_c % 2 == 1) else 0,
            "Partidos_totales": int((equipos_c * (equipos_c - 1)) / 2),
        })

    info_cats = pd.DataFrame(info_cats_rows).sort_values("Categoria").reset_index(drop=True)

    # Exportar Excel
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as w:
        # Calendario unificado
        calendario.to_excel(w, sheet_name="CALENDARIO_UNIFICADO", index=False)
        style_table(w.book["CALENDARIO_UNIFICADO"], "CALENDARIO UNIFICADO (control de rivales por fecha)")

        # Partidos por fecha
        ws_ppf = w.book.create_sheet("PARTIDOS_POR_FECHA")
        write_partidos_por_fecha_pretty(ws_ppf, cat_tables, cats)

        # Hojas por categoría (dinámico)
        for c in cats:
            sheet = safe_sheet_name(c)
            cat_tables[c].to_excel(w, sheet_name=sheet, index=False)
            style_table(w.book[sheet], f"FIXTURE {c}")

        # INFO (resumen + tabla categorías)
        info_resumen.to_excel(w, sheet_name="INFO", index=False, startrow=0)
        info_cats.to_excel(w, sheet_name="INFO", index=False, startrow=5)
        style_info_sheet(w.book["INFO"])

    print(f"🎉 Fixture generado: {OUTPUT_FILE} | Categorías detectadas: {cats}")

def generate_fixture(input_path: str, output_path: str):
    global INPUT_FILE, OUTPUT_FILE
    INPUT_FILE = input_path
    OUTPUT_FILE = output_path
    main()

if __name__ == "__main__":
    main()
